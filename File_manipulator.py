# -*- coding: utf-8 -*-
"""
Created on Tue Jul 21 09:35:37 2020

@author: Pedro Martinez
"""
from os import listdir
from os.path import isfile, join
from openpyxl import load_workbook
#import pandas as pd


class FileManipulator:
    
    def __init__(self, folderpath, filepath, extension):
        self.folderpath = folderpath
        self.filepath = filepath
        self.extension = extension
        
#        
#    folderpath = r'Z:\aie_load\A350\A350-XWB\Bricks\Site-Componentes\LGD\Multi-conf\Becario_Pedro\archivos_explorar\1024.zip'
#    #folderpath = r'Z:\aie_load\A350\A350-XWB'
#    extension = "cdf"
    folderpath = r'Z:\aie_load\A350\A350-XWB\Bricks\Site-Componentes\LGD\Multi-conf\Becario_Pedro\archivos_explorar\1024.zip'
    extension = "cdf"

    def search_files (self , folderpath=folderpath, extension=extension):
        if folderpath.endswith('.zip'):
            import zipfile
            ziper = zipfile.ZipFile(folderpath)
            all_files = ziper.namelist()
            ext_files = [f for f in all_files if f.endswith(extension)]
            
            
        else:
            import glob
            all_files = [f for f in listdir(folderpath) if isfile(join(folderpath, f))]
            
            ext_files = glob.glob(folderpath+"\\*."+extension)
            
        return all_files, ext_files


    filepath=r'Z:\aie_load\A350\A350-XWB\Bricks\Site-Componentes\LGD\Multi-conf\Becario_Pedro\1000\1000_summary_normal_cases_analysis_VCAS10.xlsx'
    
    def write_on_excel(self, contenido, filepath, hoja, columna, fila, axis=1, sufix='temp'):
        '''creates a copy of the original file, adding the values of "contenido" on the specified  sheet in "hoja". 
        axis=1 allows to write lists vertically axis=0 horizontally'''
        
        wb = load_workbook(filepath)
        sheets=wb.sheetnames
        try:
            sheet1=wb[sheets[hoja-1]]
            if isinstance(contenido, list):
                for val in contenido:    
                    sheet1.cell(row=fila,column=columna).value= val
                    if axis==1:
                           fila+=1
                    else:
                           columna+=1
    
            elif isinstance(contenido, dict):
                for key, valores in contenido.items():
                    sheet1.cell(row=fila,column=columna).value= key
                    if axis==1:
                        for val in valores:
                            sheet1.cell(row=fila+1,column=columna).value= val
                            fila+=1
                        columna+=1
                        fila-=len(valores)
                    elif axis==0:
                        for val in valores:
                            sheet1.cell(row=fila+1,column=columna).value= val
                            columna+=1
                        fila+=1
                        columna-=len(valores)
                                            
    #        elif isinstance(contenido, pd.DataFrame):
    #            with pd.ExcelWriter(filepath, engine="openpyxl", mode='a') as writer:
    #                contenido.to_excel(writer)
                    
            else:
                print('transform your object to: list, dict or DataFrame')

        except IndexError:
            print('lascagao')
            pass
    #        with pd.ExcelWriter(filepath, engine="openpyxl", mode='a') as writer:
    #            contenido.to_excel(writer,sheet_name='new_sheet')
            
            
            
        finally:
            wb.save(filepath.replace('.xlsx','')+sufix+'.xlsx')
    #        wb.close()
    
folderpath = r'Z:\aie_load\A350\A350-XWB\Bricks\Site-Componentes\LGD\Multi-conf\Becario_Pedro\archivos_explorar\1024.zip'
extension = "cdf"    
filepath=r'Z:\aie_load\A350\A350-XWB\Bricks\Site-Componentes\LGD\Multi-conf\Becario_Pedro\1000\1000_summary_normal_cases_analysis_VCAS10.xlsx'

My_object=FileManipulator(folderpath, filepath, extension)

all_files, ext_files = My_object.search_files()
My_object.write_on_excel(ext_files,filepath,1,2,9)