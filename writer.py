#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Jul 13 14:30:26 2020

@author: Pedro_Martínez
"""

from openpyxl import load_workbook
import pandas as pd
#filepath=r'/Users/macbookpro/Desktop/Python_excercises/Multiflap/1000/1000_summary_normal_cases_analysis.xlsx'

def write_on_excel(filepath,contenido, hoja, columna, fila, sufix):
    '''creates a copy of the original file, adding the values of "contenido" on the specified  sheet in "hoja". '''
    
    wb = load_workbook(filepath)
    sheets=wb.sheetnames
    try:
        sheet1=wb[sheets[hoja]]
        if isinstance(contenido, list):
            for val in contenido:    
                sheet1.cell(row=fila,column=columna).value= val
                columna+=1
        elif isinstance(contenido, dict):
            for key, valores in contenido.items():
                sheet1.cell(row=fila,column=columna).value= key
                for val in valores:
                    sheet1.cell(row=fila+1,column=columna).value= val
                    fila+=1
                columna+=1
                fila-=len(valores)
        elif isinstance(contenido, pd.DataFrame):
            with pd.ExcelWriter(filepath, engine="openpyxl", mode='a') as writer:
                contenido.to_excel(writer)
                
        else:
                
            print('transform your object to: list, dict or DataFrame')

                
    except IndexError:
        with pd.ExcelWriter(filepath, engine="openpyxl", mode='a') as writer:
            contenido.to_excel(writer,sheet_name='new_sheet')
        
        
        
    finally:
        wb.save(filepath.replace('.xlsx','')+sufix+'.xlsx')
        wb.close()
                         



#      with pd.ExcelWriter(os.path.join(path1, 'Trazabilidad350PIP.xlsx'), engine="openpyxl", mode='a') as writer:
#                    Table.to_excel(writer,sheet_name='order'+f"{pol}")